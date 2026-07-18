import re

from openpyxl import load_workbook


class AvailabilityReader:

    PERSON_HEADERS = {
        "Anestesiólogo",
        "Anestesiólogos",
        "Anesthesiologist",
    }
    WEEKDAY_HEADERS = (
        "Lunes",
        "Martes",
        "Miércoles",
        "Jueves",
        "Viernes",
    )
    SHIFT_PATTERN = re.compile(
        r"^(?P<start>\d{1,2}:\d{2})\s*(?:a|-)\s*"
        r"(?P<end>\d{1,2}:\d{2})$",
        re.IGNORECASE,
    )

    def __init__(self, filename: str):
        self.filename = filename

    def read(self):

        workbook = load_workbook(self.filename, data_only=True)
        sheet, header_row, headers = self._find_availability_sheet(
            workbook
        )
        person_column = next(
            headers[name]
            for name in self.PERSON_HEADERS
            if name in headers
        )

        availability = {}

        for row in sheet.iter_rows(
            min_row=header_row + 1,
            values_only=True,
        ):
            person = self._value(row, person_column)

            if person is None or not str(person).strip():
                continue

            person = str(person).strip()

            availability[person] = {
                weekday: self._normalize_availability(
                    self._value(row, headers[weekday])
                )
                for weekday in self.WEEKDAY_HEADERS
            }

        return availability

    def _find_availability_sheet(self, workbook):
        for sheet in workbook.worksheets:
            for row_number, row in enumerate(
                sheet.iter_rows(max_row=min(sheet.max_row, 20)),
                start=1,
            ):
                headers = {
                    str(cell.value).strip(): index
                    for index, cell in enumerate(row)
                    if cell.value is not None
                }

                has_person = bool(
                    self.PERSON_HEADERS.intersection(headers)
                )
                has_weekdays = all(
                    weekday in headers
                    for weekday in self.WEEKDAY_HEADERS
                )

                if has_person and has_weekdays:
                    return sheet, row_number, headers

        raise ValueError(
            "Availability workbook does not contain a sheet with "
            "the required person and weekday headers"
        )

    @classmethod
    def _normalize_availability(cls, value):
        if value is None:
            return None

        if not isinstance(value, str):
            return value

        text = value.strip()

        if text.casefold() == "off":
            return "OFF"

        match = cls.SHIFT_PATTERN.fullmatch(text)

        if match is None:
            return text

        start = cls._normalize_time(match.group("start"))
        end = cls._normalize_time(match.group("end"))
        return f"{start}-{end}"

    @staticmethod
    def _normalize_time(value: str) -> str:
        hour, minute = value.split(":", 1)
        return f"{int(hour):02d}:{minute}"

    @staticmethod
    def _value(row: tuple, column: int):
        return row[column] if column < len(row) else None
