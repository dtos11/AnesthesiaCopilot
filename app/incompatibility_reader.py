from openpyxl import load_workbook


class IncompatibilityReader:

    def __init__(self, filename: str):
        self.filename = filename

    def read(self):

        workbook = load_workbook(self.filename)

        sheet = workbook["Incompatibilidades"]

        incompatibilities = set()

        for row in sheet.iter_rows(min_row=2, values_only=True):

            anesthesiologist = row[0]
            surgeon = row[1]

            if not anesthesiologist or not surgeon:
                continue

            incompatibilities.add(
                (
                    anesthesiologist.strip(),
                    surgeon.strip(),
                )
            )

        return incompatibilities