from openpyxl import load_workbook


class PrivilegeReader:

    def __init__(self, filename: str):
        self.filename = filename

    def read(self):

        workbook = load_workbook(self.filename)

        privileges = {}

        for sheet in workbook.worksheets:

            if sheet.title == "README":
                continue

            people = set()

            for row in sheet.iter_rows(min_row=2, values_only=True):

                if not row[0]:
                    continue

                people.add(row[0].strip())

            privileges[sheet.title] = people

        return privileges