from pathlib import Path
import unicodedata

from openpyxl import load_workbook

from app.models.staff_alias_record import StaffAliasRecord
from app.models.staff_record import StaffRecord


class StaffDirectoryConfigurationError(ValueError):
    pass


def normalize_staff_name(name: str) -> str:
    collapsed = " ".join(str(name).strip().split())
    decomposed = unicodedata.normalize("NFKD", collapsed)
    without_diacritics = "".join(
        character
        for character in decomposed
        if not unicodedata.combining(character)
    )
    return without_diacritics.casefold()


class StaffDirectoryReader:
    REQUIRED_SHEETS = {"README", "Staff", "Aliases"}
    STAFF_HEADERS = {"HIS Full Name", "Active", "Notes"}
    ALIAS_HEADERS = {
        "HIS Full Name",
        "Alias",
        "Source",
        "Active",
        "Notes",
    }

    def __init__(self, filename: str):
        self.filename = Path(filename)

    def read(self) -> tuple[list[StaffRecord], list[StaffAliasRecord]]:
        if not self.filename.exists():
            raise FileNotFoundError(
                f"Staff directory not found: {self.filename}"
            )

        workbook = load_workbook(self.filename, data_only=True)
        missing_sheets = self.REQUIRED_SHEETS - set(workbook.sheetnames)

        if missing_sheets:
            missing = ", ".join(sorted(missing_sheets))
            raise StaffDirectoryConfigurationError(
                f"Staff directory is missing required sheets: {missing}"
            )

        staff = self._read_staff(workbook["Staff"])
        aliases = self._read_aliases(workbook["Aliases"], staff)
        self._validate_aliases(staff, aliases)

        return staff, aliases

    def _read_staff(self, sheet) -> list[StaffRecord]:
        headers = self._headers(sheet, self.STAFF_HEADERS)
        records = []

        for row_number, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            values = self._row_values(headers, row)
            his_full_name = self._text(values["HIS Full Name"])

            if his_full_name is None:
                continue

            records.append(
                StaffRecord(
                    his_full_name=his_full_name,
                    active=self._active(
                        values["Active"],
                        sheet.title,
                        row_number,
                    ),
                    notes=self._text(values["Notes"]),
                    sheet=sheet.title,
                    row=row_number,
                )
            )

        canonical_by_key = {}

        for record in records:
            key = normalize_staff_name(record.his_full_name)

            if key in canonical_by_key:
                other = canonical_by_key[key]
                raise StaffDirectoryConfigurationError(
                    f"Duplicate HIS name '{record.his_full_name}' at "
                    f"{record.sheet} rows {other.row} and {record.row}"
                )

            canonical_by_key[key] = record

        return records

    def _read_aliases(
        self,
        sheet,
        staff: list[StaffRecord],
    ) -> list[StaffAliasRecord]:
        headers = self._headers(sheet, self.ALIAS_HEADERS)
        staff_by_key = {
            normalize_staff_name(record.his_full_name): record
            for record in staff
        }
        records = []

        for row_number, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            values = self._row_values(headers, row)
            his_reference = self._text(values["HIS Full Name"])
            alias = self._text(values["Alias"])

            if his_reference is None and alias is None:
                continue

            if his_reference is None or alias is None:
                raise StaffDirectoryConfigurationError(
                    f"Alias row {row_number} must contain both "
                    "HIS Full Name and Alias"
                )

            staff_record = staff_by_key.get(
                normalize_staff_name(his_reference)
            )

            if staff_record is None:
                raise StaffDirectoryConfigurationError(
                    f"Alias '{alias}' at {sheet.title} row {row_number} "
                    f"references unknown HIS name '{his_reference}'"
                )

            records.append(
                StaffAliasRecord(
                    his_full_name=staff_record.his_full_name,
                    alias=alias,
                    source=self._text(values["Source"]),
                    active=self._active(
                        values["Active"],
                        sheet.title,
                        row_number,
                    ),
                    notes=self._text(values["Notes"]),
                    sheet=sheet.title,
                    row=row_number,
                )
            )

        return records

    def _validate_aliases(
        self,
        staff: list[StaffRecord],
        aliases: list[StaffAliasRecord],
    ) -> None:
        staff_by_key = {
            normalize_staff_name(record.his_full_name): record
            for record in staff
        }
        alias_by_key = {}

        for alias in aliases:
            key = normalize_staff_name(alias.alias)
            canonical_key = normalize_staff_name(alias.his_full_name)
            canonical_collision = staff_by_key.get(key)

            if (
                canonical_collision is not None
                and normalize_staff_name(
                    canonical_collision.his_full_name
                ) != canonical_key
            ):
                raise StaffDirectoryConfigurationError(
                    f"Alias '{alias.alias}' at {alias.sheet} row "
                    f"{alias.row} collides with HIS name "
                    f"'{canonical_collision.his_full_name}'"
                )

            if key in alias_by_key:
                other = alias_by_key[key]

                if other.his_full_name != alias.his_full_name:
                    raise StaffDirectoryConfigurationError(
                        f"Alias '{alias.alias}' maps to both "
                        f"'{other.his_full_name}' (row {other.row}) and "
                        f"'{alias.his_full_name}' (row {alias.row})"
                    )

                raise StaffDirectoryConfigurationError(
                    f"Duplicate alias '{alias.alias}' at "
                    f"{alias.sheet} rows {other.row} and {alias.row}"
                )

            alias_by_key[key] = alias

    @staticmethod
    def _headers(sheet, required: set[str]) -> dict[str, int]:
        headers = {
            str(cell.value).strip(): index
            for index, cell in enumerate(sheet[1])
            if cell.value is not None
        }
        missing = required - set(headers)

        if missing:
            names = ", ".join(sorted(missing))
            raise StaffDirectoryConfigurationError(
                f"Sheet '{sheet.title}' is missing headers: {names}"
            )

        return headers

    @staticmethod
    def _row_values(headers: dict[str, int], row: tuple) -> dict:
        return {
            header: row[index] if index < len(row) else None
            for header, index in headers.items()
        }

    @staticmethod
    def _text(value) -> str | None:
        if value is None:
            return None

        text = str(value).strip()
        return text or None

    @staticmethod
    def _active(value, sheet: str, row: int) -> bool:
        if value is None:
            return True

        if isinstance(value, bool):
            return value

        normalized = str(value).strip().casefold()
        active_values = {"yes", "true", "1", "active"}
        inactive_values = {"no", "false", "0", "inactive"}

        if normalized in active_values:
            return True

        if normalized in inactive_values:
            return False

        raise StaffDirectoryConfigurationError(
            f"Invalid Active value '{value}' at {sheet} row {row}"
        )
