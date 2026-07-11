from pathlib import Path
from openpyxl import load_workbook


class WorkbookReader:
    """
    Reads and exposes information from the daily operating room workbook.
    """

    def __init__(self, filename: str):
        self.filename = Path(filename)
        self.workbook = None
        self.sheet_names = []

    def read(self):
        """Load the Excel workbook into memory."""

        if not self.filename.exists():
            raise FileNotFoundError(f"Workbook not found: {self.filename}")

        self.workbook = load_workbook(self.filename)
        self.sheet_names = self.workbook.sheetnames

    def print_summary(self):
        """Print a simple summary of the workbook."""

        print("\nWorkbook loaded successfully.\n")

        print("Worksheets:")
        for sheet in self.sheet_names:
            print(f"  • {sheet}")